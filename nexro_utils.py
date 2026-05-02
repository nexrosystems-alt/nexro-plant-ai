"""
Nexro Plant AI - Utilidades
Exportación Excel/PDF, historial, validaciones
"""
import os
import json
import datetime
from pathlib import Path


# ═══════════════════════════════════════════════════════════════════════════════
# HISTORIAL DE ANÁLISIS
# ═══════════════════════════════════════════════════════════════════════════════
HISTORIAL_PATH = Path(os.path.dirname(os.path.abspath(__file__))) / "historial.json"
MAX_HISTORIAL = 15


def cargar_historial():
    if HISTORIAL_PATH.exists():
        try:
            with open(HISTORIAL_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []


def guardar_historial(item):
    h = cargar_historial()
    h.insert(0, item)
    h = h[:MAX_HISTORIAL]
    try:
        with open(HISTORIAL_PATH, "w", encoding="utf-8") as f:
            json.dump(h, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def limpiar_historial():
    try:
        if HISTORIAL_PATH.exists():
            HISTORIAL_PATH.unlink()
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
# VALIDACIÓN DE IMAGEN
# ═══════════════════════════════════════════════════════════════════════════════
def validar_imagen(path):
    """
    Valida una imagen antes del análisis.
    Retorna (es_valida, mensaje, severidad)
    severidad: 'ok', 'warning', 'error'
    """
    try:
        from PIL import Image
        img = Image.open(path)
        w, h = img.size

        # Muy pequeña
        if w < 100 or h < 100:
            return False, f"La imagen es muy pequeña ({w}×{h} px). Mínimo 100×100 px.", "error"

        # Muy pequeña para buen resultado
        if w < 224 or h < 224:
            return True, f"Imagen pequeña ({w}×{h} px). El resultado podría ser menos preciso.", "warning"

        # Muy grande (lento)
        if w > 4000 or h > 4000:
            return True, f"Imagen muy grande ({w}×{h} px). El análisis puede tardar un poco.", "warning"

        return True, f"Imagen óptima ({w}×{h} px).", "ok"
    except Exception as e:
        return False, f"Error al leer la imagen: {e}", "error"


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORTAR REPORTE DE ANÁLISIS A EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
def exportar_reporte_excel(path, clase, conf, info, top5_nombres, top5_confs,
                           imagen_path=None, INFO=None):
    """
    Genera un Excel profesional con el análisis completo.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Diagnostico"

    def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
    def font(bold=False, size=11, color="000000", name="Calibri"):
        return Font(name=name, bold=bold, size=size, color=color)
    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def border(color="D1D5DB"):
        side = Side(border_style="thin", color=color)
        return Border(left=side, right=side, top=side, bottom=side)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70

    # ── Header principal ──────────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    ws["A1"] = "🌿  NEXRO PLANT AI"
    ws["A1"].font = font(bold=True, size=20, color="FFFFFF")
    ws["A1"].fill = fill("0a0d14")
    ws["A1"].alignment = align("center")
    ws.row_dimensions[1].height = 48

    ws.merge_cells("A2:B2")
    ws["A2"] = "REPORTE DE DIAGNÓSTICO FITOSANITARIO"
    ws["A2"].font = font(bold=True, size=12, color="00e5a0")
    ws["A2"].fill = fill("111520")
    ws["A2"].alignment = align("center")
    ws.row_dimensions[2].height = 24

    ws.merge_cells("A3:B3")
    ws["A3"] = f"Generado el {datetime.datetime.now().strftime('%d/%m/%Y a las %H:%M')}  |  Nexro Systems"
    ws["A3"].font = font(size=10, color="7a8599", name="Calibri")
    ws["A3"].fill = fill("161b2a")
    ws["A3"].alignment = align("center")
    ws.row_dimensions[3].height = 20

    # ── Sección 1: Resumen ejecutivo ──────────────────────────────────────────
    row = 5
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "📋  RESUMEN EJECUTIVO"
    ws[f"A{row}"].font = font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = fill("00e5a0")
    ws[f"A{row}"].alignment = align("left")
    ws.row_dimensions[row].height = 28
    row += 1

    g_colors = {"ninguna":"D1FAE5","media":"FEF3C7","alta":"FED7AA","muy alta":"FECACA"}
    g_txt_colors = {"ninguna":"065F46","media":"92400E","alta":"9A3412","muy alta":"991B1B"}
    g = info.get("g", "media")
    g_bg = g_colors.get(g, "F3F4F6")
    g_txt = g_txt_colors.get(g, "374151")

    datos = [
        ("Enfermedad identificada",  f"{info.get('e','🌿')} {info.get('n', clase)}", "F9FAFB"),
        ("Clase técnica del modelo", clase, "F9FAFB"),
        ("Cultivo afectado",         _get_cultivo(clase), "F9FAFB"),
        ("Nivel de confianza",       f"{conf*100:.2f}%", "F9FAFB"),
        ("Gravedad",                 info.get("g", "").upper(), g_bg),
        ("Fecha de análisis",        datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "F9FAFB"),
    ]

    for label, value, bg in datos:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = font(bold=True, size=10, color="374151")
        ws[f"A{row}"].fill = fill("F3F4F6")
        ws[f"A{row}"].alignment = align("left", wrap=True)
        ws[f"A{row}"].border = border()

        ws[f"B{row}"] = value
        col = g_txt if label == "Gravedad" else "111827"
        ws[f"B{row}"].font = font(size=10, color=col, bold=(label=="Gravedad"))
        ws[f"B{row}"].fill = fill(bg)
        ws[f"B{row}"].alignment = align("left", wrap=True)
        ws[f"B{row}"].border = border()
        ws.row_dimensions[row].height = 24
        row += 1

    # ── Sección 2: Diagnóstico detallado ──────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "🔬  DIAGNÓSTICO DETALLADO"
    ws[f"A{row}"].font = font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = fill("00e5a0")
    ws[f"A{row}"].alignment = align("left")
    ws.row_dimensions[row].height = 28
    row += 1

    secciones = [
        ("📖 Descripción",          info.get("d", ""), "EFF6FF", "1E40AF"),
        ("💊 Tratamiento recomendado", info.get("t", ""), "FEF2F2", "991B1B"),
        ("🛡️ Medidas de prevención",  info.get("p", ""), "ECFDF5", "065F46"),
    ]

    for label, value, bg, txtcol in secciones:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = font(bold=True, size=11, color=txtcol)
        ws[f"A{row}"].fill = fill(bg)
        ws[f"A{row}"].alignment = align("left", "top", wrap=True)
        ws[f"A{row}"].border = border()

        ws[f"B{row}"] = value
        ws[f"B{row}"].font = font(size=10, color="1F2937")
        ws[f"B{row}"].fill = fill("FFFFFF")
        ws[f"B{row}"].alignment = align("left", "top", wrap=True)
        ws[f"B{row}"].border = border()
        ws.row_dimensions[row].height = max(42, min(80, len(value) // 2))
        row += 1

    # ── Sección 3: Otras posibilidades ────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "🔍  OTRAS POSIBILIDADES DETECTADAS (Top 5)"
    ws[f"A{row}"].font = font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = fill("00e5a0")
    ws[f"A{row}"].alignment = align("left")
    ws.row_dimensions[row].height = 28
    row += 1

    # Cabecera otras
    ws[f"A{row}"] = "Posibilidad"
    ws[f"A{row}"].font = font(bold=True, size=10, color="374151")
    ws[f"A{row}"].fill = fill("F3F4F6")
    ws[f"A{row}"].alignment = align("left")
    ws[f"A{row}"].border = border()

    ws[f"B{row}"] = "Confianza"
    ws[f"B{row}"].font = font(bold=True, size=10, color="374151")
    ws[f"B{row}"].fill = fill("F3F4F6")
    ws[f"B{row}"].alignment = align("center")
    ws[f"B{row}"].border = border()
    row += 1

    for i, (nm, cf) in enumerate(zip(top5_nombres, top5_confs), 1):
        inf2 = INFO.get(nm, {"n": nm, "e": "🌿"}) if INFO else {"n": nm, "e": "🌿"}
        label_n = f"#{i}  {inf2.get('e','🌿')}  {inf2.get('n', nm)}"
        bg_row = "F9FAFB" if i % 2 == 0 else "FFFFFF"

        ws[f"A{row}"] = label_n
        ws[f"A{row}"].font = font(size=10, color="111827", bold=(i==1))
        ws[f"A{row}"].fill = fill(bg_row)
        ws[f"A{row}"].alignment = align("left")
        ws[f"A{row}"].border = border()

        ws[f"B{row}"] = f"{cf*100:.2f}%"
        ws[f"B{row}"].font = font(size=10, color="00A080" if i==1 else "6B7280", bold=True)
        ws[f"B{row}"].fill = fill(bg_row)
        ws[f"B{row}"].alignment = align("center")
        ws[f"B{row}"].border = border()
        ws.row_dimensions[row].height = 22
        row += 1

    # ── Sección 4: Imagen ─────────────────────────────────────────────────────
    if imagen_path and os.path.exists(imagen_path):
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = "📷  IMAGEN ANALIZADA"
        ws[f"A{row}"].font = font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = fill("00e5a0")
        ws[f"A{row}"].alignment = align("left")
        ws.row_dimensions[row].height = 28
        row += 1

        try:
            img_xl = XLImage(imagen_path)
            # Escalar proporcional
            max_w, max_h = 400, 400
            w0, h0 = img_xl.width, img_xl.height
            ratio = min(max_w / w0, max_h / h0)
            img_xl.width  = int(w0 * ratio)
            img_xl.height = int(h0 * ratio)
            ws.add_image(img_xl, f"A{row}")
            ws.row_dimensions[row].height = img_xl.height * 0.76
            row += 1
        except Exception:
            pass

    # ── Footer ────────────────────────────────────────────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "Nexro Systems © 2026  |  nexrosystems@gmail.com  |  +57 321 521 7396"
    ws[f"A{row}"].font = font(size=9, color="6B7280", name="Calibri")
    ws[f"A{row}"].fill = fill("F9FAFB")
    ws[f"A{row}"].alignment = align("center")
    ws.row_dimensions[row].height = 22

    row += 1
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "Modelo: YOLOv8  |  88 clases  |  20 cultivos  |  Precisión: 93.7%  |  79,000+ imágenes reales"
    ws[f"A{row}"].font = font(size=8, color="9CA3AF", name="Calibri")
    ws[f"A{row}"].fill = fill("F9FAFB")
    ws[f"A{row}"].alignment = align("center")
    ws.row_dimensions[row].height = 18

    wb.save(path)
    return path


def _get_cultivo(clase):
    """Obtiene el cultivo de una clase"""
    from nexro_data import CLASE_A_CULTIVO
    return CLASE_A_CULTIVO.get(clase, "—")


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORTAR CATÁLOGO COMPLETO A EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
def exportar_catalogo_excel(path, INFO, CULTIVOS, CLASE_A_CULTIVO):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # Hoja 1: Resumen por cultivo
    ws1 = wb.active
    ws1.title = "Por Cultivo"

    def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
    def font(bold=False, size=11, color="000000"):
        return Font(name="Calibri", bold=bold, size=size, color=color)

    ws1.column_dimensions["A"].width = 4
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 50

    ws1.merge_cells("A1:D1")
    ws1["A1"] = "🌿 NEXRO PLANT AI — CATÁLOGO DE CULTIVOS Y ENFERMEDADES"
    ws1["A1"].font = font(bold=True, size=16, color="FFFFFF")
    ws1["A1"].fill = fill("0a0d14")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    headers = ["", "Cultivo", "N° Condiciones", "Enfermedades"]
    for i, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=i, value=h)
        cell.font = font(bold=True, size=11, color="FFFFFF")
        cell.fill = fill("00e5a0")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[3].height = 28

    r = 4
    for nombre, data in CULTIVOS.items():
        enfs = [INFO[c]["n"] for c in data["cls"] if c in INFO]
        enfs_str = " · ".join(enfs)
        row_bg = "F9FAFB" if r % 2 == 0 else "FFFFFF"

        ws1.cell(row=r, column=1, value=data["e"]).fill = fill(row_bg)
        ws1.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws1.cell(row=r, column=1).font = font(size=14)

        ws1.cell(row=r, column=2, value=nombre).fill = fill(row_bg)
        ws1.cell(row=r, column=2).font = font(bold=True, size=11, color="111827")

        ws1.cell(row=r, column=3, value=len(data["cls"])).fill = fill(row_bg)
        ws1.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        ws1.cell(row=r, column=3).font = font(bold=True, size=11, color="00A080")

        ws1.cell(row=r, column=4, value=enfs_str).fill = fill(row_bg)
        ws1.cell(row=r, column=4).alignment = Alignment(horizontal="left", wrap_text=True)
        ws1.cell(row=r, column=4).font = font(size=10, color="374151")
        ws1.row_dimensions[r].height = 26
        r += 1

    # Hoja 2: Todas las enfermedades
    ws2 = wb.create_sheet("Enfermedades Completo")
    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 32
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 45
    ws2.column_dimensions["F"].width = 45
    ws2.column_dimensions["G"].width = 45

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "🌿 BASE DE DATOS COMPLETA DE ENFERMEDADES"
    ws2["A1"].font = font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = fill("0a0d14")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 32

    headers2 = ["", "Enfermedad", "Cultivo", "Gravedad", "Descripción", "Tratamiento", "Prevención"]
    for i, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=i, value=h)
        cell.font = font(bold=True, size=11, color="FFFFFF")
        cell.fill = fill("00e5a0")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[3].height = 28

    g_colors = {"ninguna":"D1FAE5","media":"FEF3C7","alta":"FED7AA","muy alta":"FECACA"}
    g_txt    = {"ninguna":"065F46","media":"92400E","alta":"9A3412","muy alta":"991B1B"}

    r = 4
    for cls, inf in INFO.items():
        row_bg = "F9FAFB" if r % 2 == 0 else "FFFFFF"
        g = inf.get("g", "media")

        ws2.cell(row=r, column=1, value=inf.get("e","🌿")).fill = fill(row_bg)
        ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws2.cell(row=r, column=1).font = font(size=14)

        ws2.cell(row=r, column=2, value=inf.get("n", cls)).fill = fill(row_bg)
        ws2.cell(row=r, column=2).font = font(bold=True, size=10, color="111827")

        ws2.cell(row=r, column=3, value=CLASE_A_CULTIVO.get(cls, "—")).fill = fill(row_bg)
        ws2.cell(row=r, column=3).font = font(size=10, color="374151")

        ws2.cell(row=r, column=4, value=g.upper()).fill = fill(g_colors.get(g, "F3F4F6"))
        ws2.cell(row=r, column=4).font = font(bold=True, size=10, color=g_txt.get(g, "374151"))
        ws2.cell(row=r, column=4).alignment = Alignment(horizontal="center")

        ws2.cell(row=r, column=5, value=inf.get("d","")).fill = fill(row_bg)
        ws2.cell(row=r, column=5).font = font(size=9, color="374151")
        ws2.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")

        ws2.cell(row=r, column=6, value=inf.get("t","")).fill = fill(row_bg)
        ws2.cell(row=r, column=6).font = font(size=9, color="374151")
        ws2.cell(row=r, column=6).alignment = Alignment(wrap_text=True, vertical="top")

        ws2.cell(row=r, column=7, value=inf.get("p","")).fill = fill(row_bg)
        ws2.cell(row=r, column=7).font = font(size=9, color="374151")
        ws2.cell(row=r, column=7).alignment = Alignment(wrap_text=True, vertical="top")

        ws2.row_dimensions[r].height = 42
        r += 1

    wb.save(path)
    return path


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORTAR REPORTE PDF
# ═══════════════════════════════════════════════════════════════════════════════
def exportar_reporte_pdf(path, clase, conf, info, top5_nombres, top5_confs,
                          imagen_path=None, INFO=None):
    """Genera un PDF profesional del análisis."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Image as RLImage,
                                     Table, TableStyle, PageBreak)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY

    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = []

    # Estilos custom
    h_title = ParagraphStyle('HTitle', parent=styles['Title'], fontSize=22,
                              textColor=colors.HexColor('#00A080'), spaceAfter=6,
                              alignment=TA_CENTER, fontName='Helvetica-Bold')
    h_sub = ParagraphStyle('HSub', parent=styles['Normal'], fontSize=10,
                            textColor=colors.HexColor('#6B7280'), spaceAfter=12,
                            alignment=TA_CENTER)
    h_section = ParagraphStyle('HSection', parent=styles['Heading2'], fontSize=13,
                                textColor=colors.HexColor('#00A080'), spaceAfter=8,
                                spaceBefore=14, fontName='Helvetica-Bold')
    h_body = ParagraphStyle('HBody', parent=styles['Normal'], fontSize=10,
                             textColor=colors.HexColor('#1F2937'), alignment=TA_JUSTIFY,
                             leading=14)

    # Header
    story.append(Paragraph("🌿  NEXRO PLANT AI", h_title))
    story.append(Paragraph("Reporte de Diagnóstico Fitosanitario", h_sub))
    story.append(Paragraph(f"Generado el {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", h_sub))

    # Línea divisoria
    story.append(Spacer(1, 8))

    # Diagnóstico principal
    story.append(Paragraph("🔬  DIAGNÓSTICO PRINCIPAL", h_section))

    g_colors = {"ninguna":"#D1FAE5","media":"#FEF3C7","alta":"#FED7AA","muy alta":"#FECACA"}
    g_txt    = {"ninguna":"#065F46","media":"#92400E","alta":"#9A3412","muy alta":"#991B1B"}
    g = info.get("g", "media")

    resumen_data = [
        ["Enfermedad detectada", f"{info.get('e','🌿')}  {info.get('n', clase)}"],
        ["Clase técnica",        clase],
        ["Cultivo",              _get_cultivo(clase)],
        ["Nivel de confianza",   f"{conf*100:.2f}%"],
        ["Gravedad",             g.upper()],
    ]

    t = Table(resumen_data, colWidths=[5*cm, 11*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F3F4F6')),
        ('BACKGROUND', (1,0), (1,-1), colors.white),
        ('BACKGROUND', (1,4), (1,4), colors.HexColor(g_colors.get(g, "#F3F4F6"))),
        ('TEXTCOLOR',  (1,4), (1,4), colors.HexColor(g_txt.get(g, "#374151"))),
        ('FONTNAME',   (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME',   (1,0), (1,-1), 'Helvetica'),
        ('FONTNAME',   (1,4), (1,4), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 10),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#D1D5DB')),
        ('LEFTPADDING',(0,0), (-1,-1), 10),
        ('RIGHTPADDING',(0,0),(-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0),(-1,-1), 6),
    ]))
    story.append(t)

    # Información detallada
    story.append(Paragraph("📋  INFORMACIÓN DETALLADA", h_section))

    for icon, titulo, texto, bg, txt in [
        ("📖", "Descripción",          info.get("d",""), "#EFF6FF", "#1E40AF"),
        ("💊", "Tratamiento recomendado", info.get("t",""), "#FEF2F2", "#991B1B"),
        ("🛡️", "Medidas de prevención",  info.get("p",""), "#ECFDF5", "#065F46"),
    ]:
        t2 = Table([[f"{icon}  {titulo}", texto]], colWidths=[5*cm, 11*cm])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), colors.HexColor(bg)),
            ('TEXTCOLOR',  (0,0), (0,0), colors.HexColor(txt)),
            ('BACKGROUND', (1,0), (1,0), colors.white),
            ('FONTNAME',   (0,0), (0,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 10),
            ('VALIGN',     (0,0), (-1,-1), 'TOP'),
            ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#D1D5DB')),
            ('LEFTPADDING',(0,0), (-1,-1), 10),
            ('RIGHTPADDING',(0,0),(-1,-1), 10),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING',(0,0),(-1,-1), 8),
        ]))
        story.append(t2)
        story.append(Spacer(1, 4))

    # Otras posibilidades
    story.append(Paragraph("🔍  OTRAS POSIBILIDADES DETECTADAS", h_section))

    otras_data = [["#", "Enfermedad", "Confianza"]]
    for i, (nm, cf) in enumerate(zip(top5_nombres, top5_confs), 1):
        inf2 = INFO.get(nm, {"n": nm, "e":"🌿"}) if INFO else {"n": nm, "e":"🌿"}
        otras_data.append([
            str(i),
            f"{inf2.get('e','🌿')}  {inf2.get('n', nm)}",
            f"{cf*100:.2f}%"
        ])

    t3 = Table(otras_data, colWidths=[1.2*cm, 11.3*cm, 3.5*cm])
    t3.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#00A080')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN',      (0,0), (0,-1), 'CENTER'),
        ('ALIGN',      (2,0), (2,-1), 'CENTER'),
        ('FONTSIZE',   (0,0), (-1,-1), 10),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#D1D5DB')),
        ('LEFTPADDING',(0,0), (-1,-1), 8),
        ('RIGHTPADDING',(0,0),(-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0),(-1,-1), 6),
    ]))
    # Resaltar top 1
    t3.setStyle(TableStyle([
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (2,1), (2,1), colors.HexColor('#00A080')),
    ]))
    story.append(t3)

    # Imagen
    if imagen_path and os.path.exists(imagen_path):
        try:
            story.append(Paragraph("📷  IMAGEN ANALIZADA", h_section))
            img = RLImage(imagen_path)
            max_w = 10 * cm
            max_h = 10 * cm
            r = min(max_w / img.imageWidth, max_h / img.imageHeight)
            img.drawWidth  = img.imageWidth  * r
            img.drawHeight = img.imageHeight * r
            story.append(img)
        except Exception:
            pass

    # Footer
    story.append(Spacer(1, 20))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8,
                                   textColor=colors.HexColor('#9CA3AF'), alignment=TA_CENTER)
    story.append(Paragraph("Nexro Systems © 2026  |  nexrosystems@gmail.com  |  +57 321 521 7396", footer_style))
    story.append(Paragraph("Modelo: YOLOv8  |  88 clases  |  20 cultivos  |  Precisión: 93.7%", footer_style))

    doc.build(story)
    return path
